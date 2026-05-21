#!/usr/bin/env python3
"""Convert an Excel workbook into per-sheet HTML files (with basic style preservation)
and optionally convert those HTML files to PDF (using WeasyPrint or wkhtmltopdf).

Usage:
  python tools/convert_xlsx.py path/to/workbook.xlsx --out report/converted --pdf

This script preserves basic cell background colors, bold/italic, font color, and alignment.
"""
import argparse
import html
import os
import subprocess
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:
    raise SystemExit("Missing dependency: install openpyxl (see requirements.txt)")


def hex_from_color(color):
    if not color:
        return None
    rgb = None
    try:
        rgb = color.rgb
    except Exception:
        try:
            rgb = color.indexed
        except Exception:
            rgb = None
    if not rgb:
        return None
    # openpyxl uses ARGB like 'FF00FF00' sometimes; strip leading alpha if present
    s = str(rgb)
    if len(s) == 8:  # ARGB
        s = s[2:]
    if len(s) == 6:
        return f"#{s}"
    return None


def cell_style(cell):
    styles = []
    # background
    try:
        bg = hex_from_color(cell.fill.fgColor)
        if bg:
            styles.append(f"background:{bg}")
    except Exception:
        pass
    # font
    try:
        if cell.font and cell.font.bold:
            styles.append("font-weight:bold")
        if cell.font and cell.font.italic:
            styles.append("font-style:italic")
        if cell.font and cell.font.color:
            fc = hex_from_color(cell.font.color)
            if fc:
                styles.append(f"color:{fc}")
    except Exception:
        pass
    # alignment
    try:
        if cell.alignment and cell.alignment.horizontal:
            styles.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment and cell.alignment.vertical:
            styles.append(f"vertical-align:{cell.alignment.vertical}")
    except Exception:
        pass
    return ";".join(styles)


def sheet_to_html(sheet, max_cols=None):
    merged = {}
    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        merged[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merged[(r, c)] = None

    # determine max columns to display
    if max_cols is None:
        max_col = sheet.max_column
    else:
        max_col = max_cols

    html_lines = ["<html><head><meta charset=\"utf-8\"><style>table{border-collapse:collapse}td,th{border:1px solid #ddd;padding:4px}</style></head><body>"]
    html_lines.append(f"<h2>{html.escape(sheet.title)}</h2>")
    html_lines.append("<table>")
    for r in range(1, sheet.max_row + 1):
        html_lines.append("<tr>")
        for c in range(1, max_col + 1):
            key = (r, c)
            if key in merged and merged[key] is None:
                continue
            cell = sheet.cell(row=r, column=c)
            text = "" if cell.value is None else str(cell.value)
            text = html.escape(text)
            style = cell_style(cell)
            attrs = []
            if style:
                attrs.append(f"style=\"{style}\"")
            if key in merged and merged[key]:
                rowspan, colspan = merged[key]
                if rowspan > 1:
                    attrs.append(f"rowspan=\"{rowspan}\"")
                if colspan > 1:
                    attrs.append(f"colspan=\"{colspan}\"")
            html_lines.append(f"<td {' '.join(attrs)}>{text}</td>")
        html_lines.append("</tr>")
    html_lines.append("</table>")
    html_lines.append("</body></html>")
    return "\n".join(html_lines)


def html_to_pdf(html_path, pdf_path):
    # Try WeasyPrint first (pure python but system deps may be required)
    try:
        from weasyprint import HTML

        HTML(filename=str(html_path)).write_pdf(str(pdf_path))
        return True
    except Exception:
        pass

    # Fallback to wkhtmltopdf if available
    wk = shutil.which("wkhtmltopdf")
    if wk:
        try:
            subprocess.check_call([wk, str(html_path), str(pdf_path)])
            return True
        except Exception:
            return False
    return False


def main():
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", help="Path to workbook")
    p.add_argument("--out", default="report/converted", help="Output directory")
    p.add_argument("--pdf", action="store_true", help="Also generate PDFs (requires weasyprint or wkhtmltopdf)")
    args = p.parse_args()

    wb_path = Path(args.xlsx)
    if not wb_path.exists():
        raise SystemExit(f"Workbook not found: {wb_path}")

    out_root = Path(args.out)
    out_root.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=str(wb_path), data_only=False)
    base_name = wb_path.stem
    dest_dir = out_root / base_name
    dest_dir.mkdir(parents=True, exist_ok=True)

    for sheet in wb.worksheets:
        safe_title = sheet.title.replace('/', '_')
        html_content = sheet_to_html(sheet)
        html_file = dest_dir / f"{safe_title}.html"
        html_file.write_text(html_content, encoding="utf-8")
        print(f"Wrote: {html_file}")
        if args.pdf:
            pdf_file = dest_dir / f"{safe_title}.pdf"
            ok = html_to_pdf(html_file, pdf_file)
            if ok:
                print(f"Wrote: {pdf_file}")
            else:
                print(f"Failed to create PDF for {html_file} (weasyprint/wkhtmltopdf not available)")


if __name__ == "__main__":
    main()
